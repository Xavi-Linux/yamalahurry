"""
Module for Writer classes
"""
import abc
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Union, Tuple, TypeVar

PathLikeObj = TypeVar('PathLikeObj', str, Path)


class ConditionalTableStyle:

    pixels_per_letter: float = 1.5

    def __init__(self, anchor:str, row_header: bool = True):
        self._anchor_column: int = column_index_from_string(self._get_column_text(anchor))
        self._anchor_row: int = self._get_row_int(anchor)
        self._anchor = anchor
        self.row_header = row_header
        self.font_style: Font = Font(bold=True)
        self.row_header_alignment: Alignment = Alignment(horizontal='left')
        self.column_header_alignment: Alignment = Alignment(horizontal='center')
        self.row_header_fill: PatternFill = PatternFill(
            fill_type='solid',
            start_color='FFC0C0C0',
            end_color='FFC0C0C0'
        )
        self.column_header_fill: PatternFill = PatternFill(
            fill_type='solid',
            start_color='FFC0C0C0',
            end_color='FFC0C0C0'
        )
        self.true_format: Rule = CellIsRule(
            '==',
            ['1'],
            stopIfTrue=True,
            font=Font(color='FF3C9536'),
            fill=PatternFill(fill_type='solid', start_color='FF3C9536', end_color='FF3C9536')
        )
        self.false_format: Rule = CellIsRule(
            '==',
            ['0'],
            stopIfTrue=True,
            font=Font(color='FFFF6666'),
            fill=PatternFill(fill_type='solid', start_color='FFFF6666', end_color='FFFF6666')
        )

    @property
    def anchor(self) -> str:
        return self._anchor

    @anchor.setter
    def anchor(self, value: str) -> None:
        self._anchor_column = column_index_from_string(self._get_column_text(value))
        self._anchor_row = self._get_row_int(value)
        self._anchor = value

    def apply(self, sheet: Worksheet, max_row: int, max_column: int) -> None:
        for col_num in range(self._anchor_column, max_column + 1):
            sheet.cell(self._anchor_row, col_num).font = self.font_style
            sheet.cell(self._anchor_row, col_num).alignment = self.column_header_alignment
            sheet.cell(self._anchor_row, col_num).fill = self.column_header_fill

            #Enable column width autofit:
            sheet.column_dimensions[get_column_letter(col_num)].bestFit = True

        if self.row_header:
            for row_num in range(self._anchor_row, max_row + 1):
                sheet.cell(row_num, self._anchor_column).font = self.font_style
                sheet.cell(row_num, self._anchor_column).alignment = self.row_header_alignment
                sheet.cell(row_num, self._anchor_column).fill = self.row_header_fill

        initial_cell_ref: str = get_column_letter(self._anchor_column + 1) + str(self._anchor_row + 1)
        final_cell_ref: str = get_column_letter(max_column) + str(max_row)
        full_ref: str = initial_cell_ref + ':' + final_cell_ref
        sheet.conditional_formatting.add(full_ref, self.true_format)
        sheet.conditional_formatting.add(full_ref, self.false_format)

    @staticmethod
    def _get_row_int(cell_ref: str) -> int:
        row_substring: str = ''
        for character in cell_ref:
            try:

                if isinstance(int(character), int):
                    row_substring += character

            except ValueError:
                pass

        return int(row_substring)

    @staticmethod
    def _get_column_text(cell_ref: str) -> str:
        column_substring: str = ''
        for character in cell_ref:
            try:
                _ = int(character)

            except ValueError:
                column_substring += character

        return column_substring


class WrongInputStructure(Exception):

    def __init__(self, message: str):
        Exception.__init__(self, message)


class AbstractWriter(abc.ABC):

    def __init__(self, folderpath: PathLikeObj, *args, **kwargs):
        if isinstance(folderpath, str):
            folderpath = Path(folderpath)

        self.folderpath: Path = folderpath

    @abc.abstractmethod
    def process(self, inputs: Union[Iterable, Iterator]) -> None:
        """
        Method that receives a data structure and this used to populate an Excel Workbook
        """
        return NotImplemented()

    @abc.abstractmethod
    def save(self, filename: str) -> None:
        """
        filename will not require extension, since any saved must be .xlsx
        """
        return NotImplemented()


class OpenxlpyWriter(AbstractWriter):

    def __init__(self, folderpath: PathLikeObj):
        AbstractWriter.__init__(self, folderpath)
        self.workbook: Workbook = Workbook()
        self._input: Union[None, Dict[str, Dict[str, Union[Dict, List]]]] = None
        self._style: ConditionalTableStyle = ConditionalTableStyle(anchor='A1', row_header=True)

    def process(self, inputs: Dict[str, Dict[str, Union[Dict, List]]]) -> None:
        """
        input must have the following structure:

            {
                'sheet1_name: {
                        'rows': ['row1_text', 'row2_text' ...],
                        'columns: {
                                'column1_header': [row1_col1_value, row2_col1_value, ...],
                                'column2_header: ...
                        }
                },
                'sheet2_name: ...
            }

        """
        self._input = inputs

        self._validate_input()

        #By default, a workbook instance holds a worksheet called 'Sheet'
        self.workbook.remove(self.workbook.worksheets[0])
        for index, sheet in enumerate(self._input):
            clean_name: str = self._clear_sheet_name(sheet)
            #A sheet's name have a maximum of 31 characters:
            unique_name: str = self._generate_worksheet_name(clean_name[-31:], self.workbook.sheetnames)
            current_sheet: Worksheet = self.workbook.create_sheet(title=unique_name, index=index)
            #Build vertical axis
            for row_number, row_content in enumerate(self._input[sheet]['rows'], start=2):
                current_sheet.cell(row_number, 1, row_content)

            #Build columns' headers and content
            for col_number, header in enumerate(self._input[sheet]['columns'], start=2):
                #Header
                current_sheet.cell(1, col_number, header)
                for row, row_content in enumerate(self._input[sheet]['columns'][header], start=2):
                    current_sheet.cell(row, col_number, row_content)

            self._style.apply(current_sheet, row_number, col_number)

    def save(self, filename: str) -> None:
        filename: str = self._clear_file_name(filename)
        final_path: Path = self.folderpath / filename
        self.workbook.save(final_path.with_suffix('.xlsx'))

    def _validate_input(self) -> None:
        if isinstance(self._input, Dict):
            if len(self._input) > 0:
                if all(
                        map(
                            lambda v: isinstance(v, Dict),
                            self._input.values()
                        )
                ):
                    all_sheets_validated: bool = False
                    for sheet in self._input:
                        all_sheets_validated = False
                        try:
                            rows = self._input[sheet]['rows']
                            columns = self._input[sheet]['columns']
                            if isinstance(rows, List) and isinstance(columns, Dict):
                                rows_len: int = len(rows)
                                if all(
                                        map(
                                            lambda v: isinstance(v, List) and len(v) == rows_len,
                                            columns.values()
                                        )
                                ):
                                    all_sheets_validated = True

                        except KeyError:
                            raise WrongInputStructure(self.process.__doc__)

                    if all_sheets_validated:
                        return None

        raise WrongInputStructure(self.process.__doc__)

    @staticmethod
    def _generate_worksheet_name(name: str, current_sheets: List[str], recursion_level: int = 1) -> str:
        """
        - Avoid potential name duplications:
            if 'sheet' already exists, 'sheet' -> 'sheet_1'
            if 'sheet_1' already exists, 'sheet_1' -> 'sheet_2'
        - Proper name is found recursively
        """
        if name.lower() not in current_sheets:
            return name.lower()

        else:
            return OpenxlpyWriter._generate_worksheet_name(
                                        name.lower() + '_' + str(recursion_level),
                                        current_sheets,
                                        recursion_level + 1
            )

    @staticmethod
    def _clear_sheet_name(name: str) -> str:
        forbidden_symbols: Tuple = ('\\', '/', '*', '[', ']', ':', '?')
        replacer: str = '_'
        for symbol in forbidden_symbols:
            name = name.replace(symbol, replacer)

        return name

    @staticmethod
    def _clear_file_name(name: str) -> str:
        illegal_values: Tuple = (
                        chr(10), chr(13), '~', '"', '#', '%', '&', '*', ':', '<',
                        '>', '?', '{', '|', '}', '/','\\', '[', ']'
        )
        for value in illegal_values:
            name = name.replace(value, '_')

        return name
