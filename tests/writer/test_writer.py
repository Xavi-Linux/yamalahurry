"""
Test file for the Writer Classes
"""
import pytest
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import Dict

from yamalahurry.yamala.writer import OpenxlpyWriter, WrongInputStructure

_INPUT_ONE: Dict = {
                    'types':{
                        'rows':['squirtle', 'charmander', 'bulbasur', 'pikachu'],
                        'columns':{
                            'grass':[0, 0, 1, 0],
                            'water':[1, 0, 0, 0],
                            'fire':[0, 1, 0, 0],
                            'electric':[0, 0, 0, 1],
                            'iron':[0, 0, 0, 0]
                        }
                    }
}
_OUTPUT_ONE: Dict = {
                    'types':[
                        [2, 'A', 'squirtle'],
                        [3, 'A', 'charmander'],
                        [4, 'A', 'bulbasur'],
                        [5, 'A', 'pikachu'],
                        [1, 'B', 'grass'],
                        [1, 'C', 'water'],
                        [1, 'D', 'fire'],
                        [1, 'E', 'electric'],
                        [1, 'F', 'iron'],
                        [2, 'B', 0],
                        [3, 'B', 0],
                        [4, 'B', 1],
                        [5, 'B', 0],
                        [2, 'C', 1],
                        [3, 'C', 0],
                        [4, 'C', 0],
                        [5, 'C', 0],
                        [2, 'D', 0],
                        [3, 'D', 1],
                        [4, 'D', 0],
                        [5, 'D', 0],
                        [2, 'E', 0],
                        [3, 'E', 0],
                        [4, 'E', 0],
                        [5, 'E', 1],
                        [2, 'F', 0],
                        [3, 'F', 0],
                        [4, 'F', 0],
                        [5, 'F', 0],
                    ]
}
_INPUT_TWO: Dict = {
                    'attacks':{
                        'rows': ['fire', 'electric', 'normal'],
                        'columns': {
                            'ember':[1, 0, 0],
                            'spark':[0, 1, 0],
                            'calm mind':[0, 0, 0]
                        }
                    }
}
_OUTPUT_TWO: Dict = {
                        'attacks': [
                            [2, 'A', 'fire'],
                            [3, 'A', 'electric'],
                            [4, 'A', 'normal'],
                            [1, 'B', 'ember'],
                            [1, 'C', 'spark'],
                            [1, 'D', 'calm mind'],
                            [2, 'B', 1],
                            [3, 'B', 0],
                            [4, 'B', 0],
                            [2, 'C', 0],
                            [3, 'C', 1],
                            [4, 'C', 0],
                            [2, 'D', 0],
                            [3, 'D', 0],
                            [4, 'D', 0]
                        ]
}

_CONSOLIDATED_INPUT: Dict = _INPUT_ONE.copy()
_CONSOLIDATED_INPUT.update(_INPUT_TWO)

_CONSOLIDATED_OUTPUT: Dict = _OUTPUT_ONE.copy()
_CONSOLIDATED_OUTPUT.update(_OUTPUT_TWO)


@pytest.fixture
def make_folder(tmp_path):
    folder = tmp_path / 'tests'
    folder.mkdir()
    return folder


@pytest.fixture
def generate_writer(make_folder):
    return OpenxlpyWriter(make_folder)


# ### Happy path
@pytest.mark.parametrize(
    ('folder_path', 'expected_type'),
    [
        (#Test1
            Path('/home/path'), Path
        ),
        (#Test2
            '/home/path', Path
        )
    ], ids=[
          'Path-as-folderpath',
          'str-as-folder-path'
    ]
)
def test_instantiation(generate_writer, folder_path, expected_type):
    """
    Test to determine whether the construstor's arguments are properly parsed
    """
    writer = OpenxlpyWriter(folder_path)
    assert isinstance(writer.folderpath, expected_type)
    if isinstance(folder_path, str):
        folder_path = Path(folder_path)
    assert folder_path == writer.folderpath
    assert hasattr(writer, 'workbook')


@pytest.mark.parametrize(
    ('inputs', 'sheets_count', 'sheets_names'),
    [
        (#Test 1
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                }
            },
            1,
            ['roles']
        ),
        (#Test 2
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro\\les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 3
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro/les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 4
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro*les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 5
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro[les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 6
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro]les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 7
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro:les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 8
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro?les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            2,
            ['roles', 'ro_les']
        ),
        (#Test 9
            {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro?les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
                'ro[les':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                },
            },
            3,
            ['roles', 'ro_les', 'ro_les_1']
        ),
        (#Test 10
            {
               'b' + 'a' * 35:{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                }
            },
            1,
            ['a' * 31]
        ),
    ], ids=[
        'single-sheet-1',
        'forbidden-symbol-1',
        'forbidden-symbol-2',
        'forbidden-symbol-3',
        'forbidden-symbol-4',
        'forbidden-symbol-5',
        'forbidden-symbol-6',
        'forbidden-symbol-7',
        'two-forbidden-symbols-1',
        'a-too-large-name-for-a-sheet-1'
    ]
)
def test_excel_sheet_creation(generate_writer, inputs, sheets_count, sheets_names):
    """
    Test to determine whether the parsed content's structure is correct and the correct
    number of sheets is created. It also tests the handling of edge cases when naming sheets
    (sheet name's length and usage of forbidden symbols).
    """
    generate_writer.process(inputs)
    assert sheets_count == len(generate_writer.workbook.worksheets)
    assert sheets_names == generate_writer.workbook.sheetnames


@pytest.mark.parametrize(
    ('inputs', 'expected'),
    [
        (#Test 1
            {
                'types':{
                    'rows':['squirtle'],
                    'columns':{
                        'grass':[0]
                    }
                }
            },
            {
                'types': [
                    [2, 'A', 'squirtle'],
                    [1, 'B', 'grass'],
                    [2, 'B', 0]
                ]
            }
        ),
        (#Test 2
            _INPUT_ONE,
            _OUTPUT_ONE
        ),
        (#Test 3
            _CONSOLIDATED_INPUT,
            _CONSOLIDATED_OUTPUT
        )
    ], ids=[
        'one-sheet-one-row-one-col-1',
        'one-sheet-multiple-rows-multiple-cols-1',
        'two-sheets-multiple-rows-multiple-cols-1'
    ]
)
def test_excel_content_and_style(generate_writer, inputs, expected):
    """
    Test whether information is placed in the right cells
    """
    print(inputs)
    generate_writer.process(inputs)
    for sheet in expected:
        ws:Worksheet = generate_writer.workbook.get_sheet_by_name(sheet)
        for cells in expected[sheet]:
            cell_ref: str = cells[1] + str(cells[0])
            #Content assertion:
            assert ws[cell_ref].value == cells[2]

            #Style assertions
            if ws[cell_ref].column == 1 or ws[cell_ref].row == 1:
                assert ws[cell_ref].font.bold

                if ws[cell_ref].column == 1:
                    assert ws[cell_ref].alignment.horizontal == 'left'

                if ws[cell_ref].row == 1 and ws[cell_ref].column > 1:
                    assert ws[cell_ref].alignment.horizontal == 'center'

                assert ws[cell_ref].fill.start_color.rgb == 'FFC0C0C0'
                assert ws[cell_ref].fill.end_color.rgb =='FFC0C0C0'

            else:
                assert not ws[cell_ref].font.bold


@pytest.mark.parametrize(
    ('inputs', 'expected'),
    [
        (#Test 1
            'file',
            'file'
        ),
        (#Test 2
            'file' + chr(10),
            'file_'
        ),
        (#Test 3
            'file' + chr(13),
            'file_'
        ),
        (#Test 4
            'file' + '~',
            'file_'
        ),
        (#Test 5
            'file' + '"',
            'file_'
        ),
        (#Test 6
            'file' + '#',
            'file_'
        ),
        (#Test 7
            'file' + '%',
            'file_'
        ),
        (#Test 8
            'file' + '&',
            'file_'
        ),
        (#Test 9
            'file' + '*',
            'file_'
        ),
        (#Test 10
            'file' + ':',
            'file_'
        ),
        (#Test 11
            'file' + '<',
            'file_'
        ),
        (#Test 12
            'file' + '>',
            'file_'
        ),
        (#Test 13
            'file' + '?',
            'file_'
        ),
        (#Test 14
            'file' + '{',
            'file_'
        ),
        (#Test 15
            'file' + '|',
            'file_'
        ),
        (#Test 16
            'file' + '}',
            'file_'
        ),
        (#Test 17
            'file' + '/',
            'file_'
        ),
        (#Test 18
            'file' + '\\',
            'file_'
        ),
        (#Test 19
            'file' + '[',
            'file_'
        ),
        (#Test 4
            'file' + ']',
            'file_'
        )
    ], ids=[
        'simple-file-1',
        'forbidden-char-1',
        'forbidden-char-2',
        'forbidden-char-3',
        'forbidden-char-4',
        'forbidden-char-5',
        'forbidden-char-6',
        'forbidden-char-7',
        'forbidden-char-8',
        'forbidden-char-9',
        'forbidden-char-10',
        'forbidden-char-11',
        'forbidden-char-12',
        'forbidden-char-13',
        'forbidden-char-14',
        'forbidden-char-15',
        'forbidden-char-16',
        'forbidden-char-17',
        'forbidden-char-18',
        'forbidden-char-19'
    ]
)
def test_excel_save(generate_writer, inputs, expected):
    """
    Test if the final spreadsheet can be persisted and
    handles bad naming (use of forbidden symbols)
    """
    content = {
                'roles':{
                    'rows':['reader', 'editor', 'creator'],
                    'columns':{
                        'user1':[1, 0, 0],
                        'user2':[0, 1, 0],
                        'user3':[1, 1, 0]
                    }
                }
    }

    generate_writer.process(content)
    generate_writer.save(inputs)
    final_file = expected + '.xlsx'
    assert (generate_writer.folderpath / final_file).exists()
    assert '.xlsx' == (generate_writer.folderpath / final_file).suffix


# ### Sad path
@pytest.mark.parametrize(
    ('inputs', 'expected'),
    [
        (#Test 1
            {},
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 2
            [{}, {}],
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 3
            {
                'sheet1': {},
                'sheet2': []
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 4
            {
                'sheet1': {
                    'rows': [],
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 5
            {
                'sheet1':{
                    'columns':[],
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 6
            {
                'sheet1':{
                    'rows': 1.2,
                    'columns': {}
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 7
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns': [3, 4, 5]
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 8
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns': {
                        'column1': [1, 0],
                        'column2': [2, 0],
                        'column3': 1
                    }
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 9
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns': {
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2]
                    }
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 10
            {
                'sheet1':{
                    'rows':[1.2, 3.4, 4.5],
                    'columns': {
                        'column1':[1, 0, 3],
                        'column2':[2, 0, 3],
                        'column3':[1, 2, 3]
                    }
                },
                'sheet2': []
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 11
            {
                'sheet1':{
                    'rows':[1.2, 3.4, 4.5],
                    'columns': {
                        'column1':[1, 0, 3],
                        'column2':[2, 0, 3],
                        'column3':[1, 2, 3]
                    }
                },
                'sheet2': {
                    'rows':[1.2, 3.4, 4.5],
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 12
            {
                'sheet1':{
                    'rows':[1.2, 3.4, 4.5],
                    'columns': {
                        'column1':[1, 0, 3],
                        'column2':[2, 0, 3],
                        'column3':[1, 2, 3]
                    }
                },
                'sheet2':{
                    'columns': {
                        'column1':[1, 0, 3],
                        'column2':[2, 0, 3],
                        'column3':[1, 2, 3]
                    }
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 13
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns': {
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2, 3]
                    }
                },
                'sheet2':{
                    'rows': 1.2,
                    'columns': {
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2, 3]
                    }
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 14
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns':{
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2, 3]
                    }
                },
                'sheet2':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns': 1.2
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 15
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns':{
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2, 3]
                    }
                },
                'sheet2':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns':{
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': 'Charmander'
                    }
                }
            },
            OpenxlpyWriter.process.__doc__
        ),
        (#Test 16
            {
                'sheet1':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns':{
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2, 3]
                    }
                },
                'sheet2':{
                    'rows': [1.2, 3.4, 4.5],
                    'columns':{
                        'column1': [1, 0, 3],
                        'column2': [2, 0, 3],
                        'column3': [1, 2, 3, 5]
                    }
                }
            },
            OpenxlpyWriter.process.__doc__
        )
    ], ids=[
        'empty-dict-1',
        'list-of-dicts-1',
        'no-dict-for-first-level-values-1',
        'rows-no-columns-1',
        'columns-no-rows-1',
        'rows-value-not-a-list-1',
        'columns-value-not-a-dict-1',
        'columns-dict-values-not-lists-1',
        'column-lists-length-not-equals-rows-length-1',
        'no-dict-for-first-level-values-2',
        'rows-no-columns-2',
        'columns-no-rows-2',
        'rows-value-not-a-list-2',
        'columns-value-not-a-dict-2',
        'columns-dict-values-not-lists-2',
        'column-lists-length-not-equals-rows-length-2'
    ]
)
def test_wrong_input(generate_writer, inputs, expected):
    """
    Check whether the code raises error when a writer
    receives a wrong input data structure
    """
    with pytest.raises(WrongInputStructure) as exp:
        generate_writer.process(inputs)
    assert expected == exp.value.args[0]


#@pytest.mark.skip
def test_full_flow():
    """
    Test the full flow and generate an actual Excel file, so that it can be visually inspected
    """
    target_folder: str = '/home/xavi/Documents/Pynotes/yamalaHarris/yamalahurry/yamala/output'
    writer = OpenxlpyWriter(target_folder)
    writer.process(_CONSOLIDATED_INPUT)
    writer.save('test_7.xlsx')
    assert True
